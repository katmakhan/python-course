# Import Excels for read and write
import xlrd

import openpyxl
from openpyxl import load_workbook

import xlwings as xw

import csv
import pandas as pd
import os
import shutil

import time
import datetime
from datetime import timedelta
from datetime import datetime as dt

import json
import requests

import zipfile


# Raw string
def to_raw(string):
    return fr"{string}"

#Writing JSON
def writejson(location,jsondt):
	with open(location, 'w+', encoding='utf-8') as f:
		json.dump(jsondt, f, ensure_ascii=False, indent=4)
		print("Sucessfully written into json file")
		f.close()
	
		
#Reading JSON
def readjson(location):
	with open(location, 'r', encoding='utf-8') as f:
		data=json.load(f)
		print("Sucessfully read json file")
		return data
		f.close()

	
# Convert date into timeinmill
def convert_timein_mill(date_obj):
	#print("Converting time into mill"+str(date_obj))
	timeinmill=int(round(date_obj. timestamp()* 1000))
	return timeinmill
	
# Converting the date to string	
def convert_date_to_str(date_obj,date_format):
	time_str_local=date_obj.strftime(date_format)
	return time_str_local
	
# Converting the string into date
def convert_str_to_date(date_str,date_format):
	date_obj=dt.strptime(date_str,date_format)
	return date_obj
	
# Next 5 minute
def find_time_after_minutes(time_object,interval):
	next_date_obj=time_object+ datetime.timedelta(minutes=interval)
	return next_date_obj


# Adding only near expiry data into the json
def convert_to_required_data(full_data_local,expiry):
	required_data_local={}
	for keys in full_data_local:
		#print(full_data_local[keys]['exp'])
		if full_data_local[keys]['exp']==expiry:
			required_data_local[keys]={}			
			required_data_local[keys]=full_data_local[keys]
	return required_data_local

# Find the minutes between initial and end time
def find_minutes_between(initial_time_obj,end_time_obj):	
	time_delta = (end_time_obj - initial_time_obj)
	total_seconds = time_delta.total_seconds()
	minutes = total_seconds/60
	return int(minutes)

#Writing Excel
# def write_toexcel(location,jsondt,sheetname,filename_str):
# 	# Transponse the data
# 	if(len(jsondt)>1):
# 		df = pd.DataFrame(jsondt).T
# 		df =  df[["Stock Name","Stock code","BSE code","ISIN","Industry Name","Market Capitalization","12-DEC-2022"]]

# 		# df.to_excel(location+filename_str+'.xlsx',sheet_name=sheetname)
# 		# filepath=location+filename_str+'.xlsx'
# 		filepath = os.path.join(os.getcwd()+ "\\"+filename_str+'.xlsx')

# 		# filepath='C:\\Users\\HP\\Downloads\\Adrash\\output.xlsx'
# 		with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:  
# 				df.to_excel(writer,sheet_name=sheetname,index=False, startrow=2)
# 	else:
# 		print("\n")
# 		print("*******************************")
# 		print("ERROR: Output file is empty....")
# 		print("*******************************")
# 		print("\n")

def write_toexcel_new(jsondt,sheetname,filename_str):
	# Call a Workbook() function of openpyxl 
	# to create a new blank Workbook object
	filepath = os.path.join(os.getcwd()+ "\\"+filename_str+'.xlsx')
	# filepath=r'C:\\Users\\HP\\Downloads\\Adrash\\output.xlsx'
	# print(filepath)

	wb = openpyxl.Workbook()
	wb=load_workbook(filename = filepath)
	  
	# Get workbook active sheet  
	# from the active attribute. 
	# sheet = wb.get_active_sheet()
	# sheet = wb.active
	sheet=wb['result']

	#Clear all values existing in the excel sheet
	for row in sheet.iter_rows():
		for cell in row:
		   	cell.value = None
	  
	# writing to the specified cell
	# sheet.cell(1,1).value = "testing"

	# set the height of the row
	# sheet.row_dimensions[1].height = 70
	# set the width of the column
	# sheet.column_dimensions['B'].width = 20

	#To fill up the headers
	for row in jsondt:
		data=jsondt[row]
		# print(data)
		column=0
		for field in data:
			# print(field)
			column=column+1
			sheet.cell(row,column).value = field
			# print(field)
		break

	# Remove no entry data
	newdata={}
	count=0
	for row in jsondt:
		data=jsondt[row]
		if len(data)>6:
			count=count+1
			newdata[count]=data
			
			
	# To add values now
	for row in newdata:
		data=newdata[row]
		column=0
		for field in data:
			# print(field)
			column=column+1
			sheet.cell(row+1,column).value = data[field]
			# print(data[field])
	  
	# save the file
	wb.save(filepath)

# def write_toexcel_new(jsondt,sheetname,filename_str):
# 	print("Writing to the excel")
# 	# Call a Workbook() function of openpyxl 
# 	# to create a new blank Workbook object
# 	filepath = os.path.join(os.getcwd()+ "\\"+filename_str+'.xlsx')
# 	# filepath=r'C:\\Users\\HP\\Downloads\\Adrash\\output.xlsx'
# 	# print(filepath)
# 	wb = xw.Book(filepath)

# 	sheet = wb.sheets['result']  
# 	# writing to the specified cell
# 	# sheet.cell(1,1).value = "testing"

# 	# set the height of the row
# 	# sheet.row_dimensions[1].height = 70
# 	# set the width of the column
# 	# sheet.column_dimensions['B'].width = 20
# 	for row in jsondt:
# 		data=jsondt[row]
# 		# print(data)
# 		column=0
# 		for field in data:
# 			# print(field)
# 			column=column+1
# 			sheet.range(row,column).value = field
# 			# print(field)
# 		break
# 	print("done...")

# 	for row in jsondt:
# 		data=jsondt[row]
# 		# print(data)
# 		column=0
# 		print("writing data to row ",row)
# 		for field in data:
# 			# print(field)
# 			column=column+1
# 			sheet.range(row+1,column).value = data[field]
# 			# print("writing data into the table",data[field])
	  
# 	# save the file
# 	print("Saving...")
# 	wb.save()

# Stock Name	Stock code	BSE code	ISIN	Industry Name	Market Capitalization in Cr
def read_fromexcel():
	wbook=xlrd.open_workbook(resourcefoldername+"/"+resourcefilename)
	sheet=wbook.sheet_by_index(0)
	total_rows=sheet.nrows

	industrydata={}
	for i in range(1,total_rows-1):
		industrydata[i]={}
		industrydata[i]['Stock Name']=sheet.cell_value(i,0)
		industrydata[i]['Stock code']=sheet.cell_value(i,1)
		industrydata[i]['BSE code']=sheet.cell_value(i,2)
		industrydata[i]['ISIN']=sheet.cell_value(i,3)
		industrydata[i]['Industry Name']=sheet.cell_value(i,4)
		if not sheet.cell_value(i,5)=="":
			industrydata[i]['Market Capitalization']=float(sheet.cell_value(i,5))

	# print(optiondata)
	print("File read sucesfully")
	return industrydata
	# write_toexcel(optiondata,"newfilename.xls")

inputpath='./bhavfiles/'
outputpath='./output/'
resourcefilename="NSE_BSE Stocks.xlsx"
resourcefoldername="resource"
total_days=30
csvpath="./csvbhav"
bhav_date_format="%d%b%Y" #02082021
output_date_format="%d-%m-%Y" #02-08-2021

def get_allfiles():
	arr = os.listdir(inputpath)
	fileslist={}
	for file in arr:
		if ".csv" in file:
			# print(file)
			fileslist[file]=file

	# print(fileslist)
	return fileslist

from collections import OrderedDict

def get_allbhavfiles():
	arr = os.listdir(csvpath)
	fileslist={}
	for file in arr:
		if ".csv" in file:
			# print(file)
			fileslist[file]=file

	# print(fileslist)
	return OrderedDict(reversed(list(fileslist.items())))

# Filter banknifty and nifty seperately
def filterdata(dataset,filtername):
	output={}
	for row in dataset:
		tickname=dataset[row]['Ticker']

		if(tickname.startswith(filtername)):
			output[row]=dataset[row]

	return output

def convert_listdata_dict(dataset):
	output={}
	for row in dataset:
		name=dataset[row]['Ticker']
		time=dataset[row]['Time']
		# Full data will be under, name>time
		output[name+time]=dataset[row]

	return output

def read_fromcsv(filename):
	output_dict=pd.read_csv(filename).to_dict('index')
	return output_dict

import urllib.request
import http.cookiejar

def downloadbhav(date_str):
	# url="https://www1.nseindia.com/ArchieveSearch?h_filetype=eqbhav&date=13-12-2022&section=EQ"
	# https://www1.nseindia.com/content/historical/EQUITIES/2023/JAN/cm02JAN2023bhav.csv.zip
	# https://www1.nseindia.com/content/historical/EQUITIES/2023/JAN/cm02JAN2023bhav.csv.zip
	# https://www1.nseindia.com/content/historical/EQUITIES/2022/DEC/cm13DEC2022bhav.csv.zip

	# date_str="13DEC2022"
	# https://www1.nseindia.com/ArchieveSearch?h_filetype=eqbhav&date=02-01-2023&section=EQ

	year=date_str[-4:]
	month=date_str.split(year)[0][2:]

	#Adding a new link to bypass cookies
	s = requests.Session()

	# url="https://www1.nseindia.com/ArchieveSearch?h_filetype=eqbhav&date=02-01-2023&section=EQ"

	# headers = {
	# "Cookie": "NSE-TEST-1=1910513674.20480.0000",
	# # "Connection": "keep-alive",
	# # "Accept-Encoding": "gzip, deflate, br",
	# # "Accept-Language": "en-US,en;q=0.5",
	# # "Upgrade-Insecure-Requests": "1",
	# # "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
	# "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"}
	# res = s.get(url,headers=headers)

	#Downloading Bhav
	# https://archives.nseindia.com/content/historical/EQUITIES/2023/APR/cm13APR2023bhav.csv.zip
	url="https://archives.nseindia.com/content/historical/EQUITIES/"+str(year)+"/"+str(month)+"/cm"+date_str+"bhav.csv.zip"


	# print("Checking ",url)

	#Method 1
	try:
		res = s.get(url,timeout=2)

		#Method 2
		# response = urllib.request.urlopen(url)
		# data = response.read()
		# print(data)

		#Method 3
		# Create a cookie jar to store the cookie
		# cj = http.cookiejar.CookieJar()

		# # Create an opener to handle cookies
		# opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

		# # Send the request and retrieve the file
		# response = opener.open(url)
		# data = response.read()
		# print(data)

		# print(res.status_code)
		if(res.status_code==200):
			print("File exist...")
			open('./bhavfiles/cm'+date_str+'bhav.csv.zip', 'wb+').write(res.content)
		else:
			print("--------------------No bhav data found in ",date_str)
		# print("wait for 3 sec..")
		# time.sleep(3)
	except Exception as e:
		print("No data found..")
		# print(e)

def bhavdownloadloop():
	today_obj=dt.today()
	for i in range (total_days):
		start_date=convert_date_to_str(today_obj,bhav_date_format).upper()
		print("Getting bhav data of ",start_date)
		today_obj=today_obj-timedelta(days = 1)
		downloadbhav(start_date)
	
def unziploop():
	allfilenames=get_allfiles()
	for filename in allfilenames:
		print(filename)
		with zipfile.ZipFile("bhavfiles/"+filename,"r") as zip_ref:
			zip_ref.extractall("csvbhav")

#Stock Name	Stock code	BSE code	ISIN	Industry Name	Market Capitalization in Cr
def getresourcedata():
	industrydata=read_fromexcel()
	# print(industrydata)
	return industrydata

def updatethevalues(industrydata,singlebhavdata):
	percentchange=0
	findstock=False
	checkdate=str(singlebhavdata['TIMESTAMP'])

	for row in industrydata:
		singledata=industrydata[row]
		# print(singledata['ISIN'])
		# print(singlebhavdata['ISIN'])

		if singledata['ISIN']==singlebhavdata['ISIN']:
			# print(singledata)
			# print(singlebhavdata)
			percentchange=(singlebhavdata['CLOSE']-singlebhavdata['PREVCLOSE'])/singlebhavdata['PREVCLOSE']
			findstock=True
			# print("Percent change is ",percentchange)
			industrydata[row][checkdate]=percentchange

			#To cross check when there is some data missmatch
			# if singlebhavdata['SYMBOL']=="INE534E01020":
			# 	print(percentchange)

	if not findstock:
		# print("No ",singlebhavdata['SYMBOL']," ",singlebhavdata['ISIN'], " found in bhav ",checkdate)
		with open('datamissing.txt', 'a') as f:
			# Write the string to the file
			f.write("No "+singlebhavdata['SYMBOL']+" "+singlebhavdata['ISIN']+ " found in bhav "+checkdate+"\n")

	return industrydata
def removefield(dataset,fieldname):
	newdata=dataset
	for row in dataset:
		data=dataset[row]
		newdata[row].pop(fieldname)

	return newdata

#Order the bhav files
def orderthedata(filelist):
	datelist=[]
	for filename in filelist:
		# print(filename)
		date_str=filename.split("bhav")[0].split("cm")[1]
		# print(date_str)
		date_obj=convert_str_to_date(date_str,"%d%b%Y")
		datelist.append(date_obj)

	datelist.sort(reverse=True)
	# print(datelist)


	newlist=[]
	for date in datelist:
		date_str=convert_date_to_str(date,"%d%b%Y")
		filename="cm"+date_str+"bhav.csv"
		newlist.append(filename)

	# print(newlist)
	return newlist


def processbhavdata(industrydata):
	allfilenames=get_allbhavfiles()
	allfilenames=orderthedata(allfilenames)
	for filename in allfilenames:
		# print(filename)
		bhavjson=read_fromcsv(csvpath+"/"+filename)
		# print(bhavjson)
		processingdate=filename.split("cm")[1].split("bhav")[0]
		print("processing data of ",processingdate)
		processingdate_obj=convert_str_to_date(processingdate,bhav_date_format)
		outputdate_str=convert_date_to_str(processingdate_obj,output_date_format)
		# print(outputdate_str)

		for child in bhavjson:
			singlebhavdata=bhavjson[child]

			# print(singlebhavdata)

			#updatetheindustry data with the date and segment value change
			industrydata=updatethevalues(industrydata,singlebhavdata)
		
	# print(industrydata[1])
	industrydata=removefield(industrydata,"BSE code")
	
	#After processing all data, write into the output file.

	# location,jsondt,sheetname,filename_str
	# write_toexcel("./",industrydata,"result","output")
	write_toexcel_new(industrydata,"result","output")

def deleteoldata():
	folderpath=os.path.join(os.getcwd(),csvpath)
	try:
		shutil.rmtree(folderpath)
		print("Deleted the old files csv...")
	except Exception as e:
		# print('Failed to delete %s. Reason: %s' % (folderpath, e))
		print("No bhav csv folder found")


	folderpath=os.path.join(os.getcwd(),inputpath)	
	try:
		shutil.rmtree(folderpath)
		print("Deleted the old files zip...")
	except Exception as e:
		# print('Failed to delete %s. Reason: %s' % (folderpath, e))
		print("No bhav zip folder found")

def createfolders():
	if not os.path.exists(csvpath):
		print("Creating csv folder..")
		os.makedirs(csvpath)
	if not os.path.exists(inputpath):
		print("Creating zip folder..")
		os.makedirs(inputpath)

# Main Function
try:
	# Only once a day
	deleteoldata()
	createfolders()
	bhavdownloadloop()
	print("\n\n")
	unziploop()	
	with open('datamissing.txt', 'w') as f:
		# Write an empty string to the file
		f.write('')

	#Everytime
	industrylist=getresourcedata()
	processbhavdata(industrylist)

	print("\n\n")
	print("______________________________")
	print("All Data Completely Processed.")
	print("------------------------------")
	print("______________________________")

	# filelist=get_allbhavfiles()
	# filelist=orderthedata(filelist)
	
	
except Exception as e:
	print("*********Exception Occured in main Function**********")
	print(e)